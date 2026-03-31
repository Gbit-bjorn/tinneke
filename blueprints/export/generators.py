"""
Export Generators voor BK/DPK/LPD Attestatiesysteem.

Bevat functies voor HTML en Excel export:
  - genereer_html_attestering: Print-klare HTML voor individuele leerling
  - genereer_excel_attestering: Excel met BK/DPK/LPD rijen
  - genereer_excel_klasoverzicht: Klasoverzicht als Excel (leerlingen × BKs)
"""

from datetime import datetime
from io import BytesIO
from typing import Optional


def genereer_html_attestering(leerling: dict, doelen: list, resultaten: dict) -> str:
    """
    Genereert een print-klare HTML string voor attestering.

    Args:
        leerling: {'id': int, 'naam': str, 'voornaam': str, 'klas_id': int, ...}
        doelen: [{...}] lijst van BK/DPK/LPD doelen (hierarchie)
        resultaten: {lpd_id: True/False} of {lpd_uuid: {'behaald': bool, 'datum': str}}

    Returns:
        Standalone HTML string met inline CSS.
    """
    from datetime import date

    # Extract leerling info
    naam = f"{leerling.get('voornaam', '')} {leerling.get('naam', '')}".strip()
    klas_id = leerling.get('klas_id', '')
    schooljaar = datetime.now().year
    export_datum = date.today().strftime("%d/%m/%Y")

    # CSS inline
    css = """
body {
    font-family: 'Segoe UI', Arial, sans-serif;
    margin: 32px;
    color: #222;
    background: #fff;
}

.header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    border-bottom: 3px solid #1565C0;
    padding-bottom: 16px;
    margin-bottom: 24px;
}

.header-left h1 {
    color: #1565C0;
    font-size: 24px;
    margin: 0;
}

.header-left p {
    color: #666;
    font-size: 12px;
    margin: 4px 0 0 0;
}

.header-right {
    width: 80px;
    height: 50px;
    background: #e3f2fd;
    border: 2px dashed #90CAF9;
    border-radius: 4px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 10px;
    color: #90CAF9;
}

h2 {
    color: #1565C0;
    margin-top: 24px;
    margin-bottom: 12px;
    padding-bottom: 4px;
    border-bottom: 1px solid #e0e0e0;
    font-size: 16px;
}

h3 {
    color: #37474F;
    margin-left: 20px;
    margin-top: 12px;
    margin-bottom: 8px;
    font-size: 14px;
}

.dpk-section {
    margin-left: 20px;
    margin-bottom: 8px;
}

.lpd {
    margin-left: 40px;
    padding: 4px 0;
    font-size: 13px;
    display: flex;
    align-items: center;
}

.lpd-icon {
    font-size: 16px;
    margin-right: 8px;
    width: 20px;
    text-align: center;
}

.lpd-behaald .lpd-icon {
    color: #2E7D32;
}

.lpd-niet-behaald .lpd-icon {
    color: #C62828;
}

.status-badge {
    margin-left: 12px;
    padding: 2px 8px;
    border-radius: 3px;
    font-size: 11px;
    font-weight: bold;
}

.status-behaald {
    background: #E8F5E9;
    color: #2E7D32;
}

.status-niet-behaald {
    background: #FFEBEE;
    color: #C62828;
}

.meta {
    background: #f5f5f5;
    padding: 12px;
    border-radius: 4px;
    margin-top: 24px;
    font-size: 12px;
    color: #666;
}

@media print {
    body { margin: 0; padding: 20px; }
    .header { page-break-after: avoid; }
    h2 { page-break-after: avoid; }
    h3 { page-break-after: avoid; }
}
"""

    # Begin HTML
    html_lines = [
        "<!DOCTYPE html>",
        "<html lang='nl'>",
        "<head>",
        "  <meta charset='utf-8'>",
        "  <meta name='viewport' content='width=device-width, initial-scale=1.0'>",
        f"  <title>Attestering {naam}</title>",
        "  <style>",
        css,
        "  </style>",
        "</head>",
        "<body>",
        "  <div class='header'>",
        "    <div class='header-left'>",
        f"      <h1>Attestering BK/DPK/LPD</h1>",
        f"      <p><b>{naam}</b> | Klas {klas_id} | Schooljaar {schooljaar - 1}-{schooljaar}</p>",
        "    </div>",
        "    <div class='header-right'>LOGO</div>",
        "  </div>",
    ]

    # Organiseer doelen in BK > DPK > LPD hiërarchie
    # Doelen kunnen van twee bronnen komen:
    # 1. LLinkid API: { 'key', 'type', 'titel', 'nr', 'depth', ... }
    # 2. Lokale DB: { 'bk_code', 'bk_naam', 'dpk_code', ..., 'lpd_code', ... }

    bks_dict = {}
    for doel in doelen:
        # Probeer beide formats
        bk_code = doel.get('bk_code', doel.get('code', doel.get('nr', 'BK?')))
        bk_naam = doel.get('bk_naam', doel.get('naam', doel.get('titel', 'Onbekend')))
        dpk_code = doel.get('dpk_code', doel.get('nr', 'DPK?'))
        dpk_naam = doel.get('dpk_naam', doel.get('titel', 'Onbekend'))
        lpd_id = doel.get('lpd_id', doel.get('id', doel.get('key', '')))
        lpd_code = doel.get('lpd_code', doel.get('code', doel.get('nr', 'LPD?')))
        lpd_omsch = doel.get('lpd_omschrijving', doel.get('omschrijving', doel.get('titel', '')))

        if bk_code not in bks_dict:
            bks_dict[bk_code] = {'naam': bk_naam, 'dpks': {}}

        if dpk_code not in bks_dict[bk_code]['dpks']:
            bks_dict[bk_code]['dpks'][dpk_code] = {'naam': dpk_naam, 'lpds': []}

        bks_dict[bk_code]['dpks'][dpk_code]['lpds'].append({
            'id': lpd_id,
            'code': lpd_code,
            'omschrijving': lpd_omsch
        })

    # Rendeer BK > DPK > LPD
    for bk_code, bk_data in sorted(bks_dict.items()):
        # Bereken BK status
        all_lpds = []
        for dpk_data in bk_data['dpks'].values():
            all_lpds.extend(dpk_data['lpds'])

        bk_behaald = all(resultaten.get(lpd['id'], False) for lpd in all_lpds) if all_lpds else False
        bk_status = "BEHAALD ✓" if bk_behaald else "NIET BEHAALD"
        status_cls = "status-behaald" if bk_behaald else "status-niet-behaald"

        html_lines.append(f"  <h2>{bk_code}: {bk_data['naam']}")
        html_lines.append(f"    <span class='status-badge {status_cls}'>{bk_status}</span>")
        html_lines.append("  </h2>")

        # DPKs
        for dpk_code, dpk_data in sorted(bk_data['dpks'].items()):
            dpk_behaald = all(resultaten.get(lpd['id'], False) for lpd in dpk_data['lpds'])
            dpk_status = "BEHAALD ✓" if dpk_behaald else "NIET BEHAALD"
            dpk_status_cls = "status-behaald" if dpk_behaald else "status-niet-behaald"

            html_lines.append(f"  <h3>{dpk_code}: {dpk_data['naam']}")
            html_lines.append(f"    <span class='status-badge {dpk_status_cls}'>{dpk_status}</span>")
            html_lines.append("  </h3>")

            # LPDs
            html_lines.append("  <div class='dpk-section'>")
            for lpd in dpk_data['lpds']:
                behaald = resultaten.get(lpd['id'], False)
                icon = "✓" if behaald else "○"
                cls = "lpd-behaald" if behaald else "lpd-niet-behaald"
                html_lines.append(f"    <div class='lpd {cls}'>")
                html_lines.append(f"      <span class='lpd-icon'>{icon}</span>")
                html_lines.append(f"      <span><b>{lpd['code']}</b>: {lpd['omschrijving']}</span>")
                html_lines.append("    </div>")
            html_lines.append("  </div>")

    # Footer
    html_lines.append(f"  <div class='meta'>")
    html_lines.append(f"    <strong>Exportdatum:</strong> {export_datum} | ")
    html_lines.append(f"    <strong>Schooljaar:</strong> {schooljaar - 1}-{schooljaar}")
    html_lines.append(f"  </div>")
    html_lines.append("</body>")
    html_lines.append("</html>")

    return "\n".join(html_lines)


def genereer_excel_attestering(leerling: dict, doelen: list, resultaten: dict) -> bytes:
    """
    Genereert Excel bestand voor attestering (openpyxl).

    Args:
        leerling: Student info dict
        doelen: List of goals/objectives
        resultaten: Dict met behaalde statussen

    Returns:
        Bytes van het Excel-bestand.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        # Fallback: return CSV bytes
        return _genereer_csv_bytes(leerling, doelen, resultaten)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attestering"

    # Styling
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    behaald_fill = PatternFill("solid", fgColor="C8E6C9")
    behaald_font = Font(color="1B5E20", bold=True)
    niet_behaald_fill = PatternFill("solid", fgColor="FFCDD2")
    niet_behaald_font = Font(color="B71C1C")
    alt_fill = PatternFill("solid", fgColor="F5F5F5")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Header info
    naam = f"{leerling.get('voornaam', '')} {leerling.get('naam', '')}".strip()
    klas = leerling.get('klas_id', '')
    schooljaar = datetime.now().year

    ws.append([f"Attestering: {naam} | Klas: {klas} | Schooljaar: {schooljaar - 1}-{schooljaar}"])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    # Column headers
    headers = ["BK/DPK/LPD Code", "Omschrijving", "Soort", "Behaald", "Datum"]
    ws.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Organize data into hierarchy
    bks_dict = {}
    for doel in doelen:
        bk_code = doel.get('bk_code', doel.get('code', doel.get('nr', 'BK?')))
        bk_naam = doel.get('bk_naam', doel.get('naam', doel.get('titel', 'Onbekend')))
        dpk_code = doel.get('dpk_code', doel.get('nr', 'DPK?'))
        dpk_naam = doel.get('dpk_naam', doel.get('titel', 'Onbekend'))
        lpd_id = doel.get('lpd_id', doel.get('id', doel.get('key', '')))
        lpd_code = doel.get('lpd_code', doel.get('code', doel.get('nr', 'LPD?')))
        lpd_omsch = doel.get('lpd_omschrijving', doel.get('omschrijving', doel.get('titel', '')))

        if bk_code not in bks_dict:
            bks_dict[bk_code] = {'naam': bk_naam, 'dpks': {}}
        if dpk_code not in bks_dict[bk_code]['dpks']:
            bks_dict[bk_code]['dpks'][dpk_code] = {'naam': dpk_naam, 'lpds': []}

        bks_dict[bk_code]['dpks'][dpk_code]['lpds'].append({
            'id': lpd_id, 'code': lpd_code, 'omschrijving': lpd_omsch
        })

    # Write rows
    row_num = 4
    is_even = False

    for bk_code in sorted(bks_dict.keys()):
        bk_data = bks_dict[bk_code]

        # BK row
        ws.append([bk_code, bk_data['naam'], "BK", "", ""])
        bk_row = ws[row_num]
        for cell in bk_row:
            cell.font = Font(bold=True, size=10)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor="E3F2FD")
        row_num += 1

        for dpk_code in sorted(bk_data['dpks'].keys()):
            dpk_data = bk_data['dpks'][dpk_code]

            # DPK row
            ws.append(["  " + dpk_code, dpk_data['naam'], "DPK", "", ""])
            dpk_row = ws[row_num]
            for cell in dpk_row:
                cell.font = Font(bold=True, size=9)
                cell.border = border
            row_num += 1

            # LPD rows
            for lpd in dpk_data['lpds']:
                behaald = resultaten.get(lpd['id'], False)
                behaald_text = "JA" if behaald else "NEE"
                datum_text = ""

                ws.append([
                    "    " + lpd['code'],
                    lpd['omschrijving'],
                    "LPD",
                    behaald_text,
                    datum_text
                ])

                lpd_row = ws[row_num]

                # Apply alternating background
                if is_even:
                    for cell in lpd_row:
                        cell.fill = alt_fill

                # Format cells
                lpd_row[1].alignment = left_align
                lpd_row[4].alignment = center_align

                # Behaald column formatting
                behaald_cell = lpd_row[4]
                if behaald:
                    behaald_cell.fill = behaald_fill
                    behaald_cell.font = behaald_font
                else:
                    behaald_cell.fill = niet_behaald_fill
                    behaald_cell.font = niet_behaald_font

                for cell in lpd_row:
                    cell.border = border
                    cell.alignment = center_align

                row_num += 1
                is_even = not is_even

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def genereer_excel_klasoverzicht(klas: dict, leerlingen: list, stats: list) -> bytes:
    """
    Genereert Excel klasoverzicht (leerlingen als rijen, BKs als kolommen).

    Args:
        klas: {'id': int, 'naam': str, 'schooljaar': int}
        leerlingen: List of leerling dicts
        stats: List of per-leerling completion stats per BK

    Returns:
        Bytes van het Excel-bestand.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return _genereer_csv_bytes_klas(klas, leerlingen, stats)

    wb = Workbook()
    ws = wb.active
    ws.title = "Klasoverzicht"

    # Styling
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    behaald_fill = PatternFill("solid", fgColor="C8E6C9")
    behaald_font = Font(color="1B5E20", bold=True)
    gedeeltelijk_fill = PatternFill("solid", fgColor="FFE0B2")
    niet_behaald_fill = PatternFill("solid", fgColor="FFCDD2")
    alt_fill = PatternFill("solid", fgColor="F9F9F9")

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Meta info
    klas_naam = klas.get('naam', f"Klas {klas.get('id')}")
    schooljaar = klas.get('schooljaar', datetime.now().year)

    ws.append([f"Klasoverzicht: {klas_naam} | Schooljaar: {schooljaar - 1}-{schooljaar}"])
    ws.merge_cells("A1:Z1")
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])

    # Extract unique BK codes from stats
    bk_codes = set()
    for stat in stats:
        bk_codes.update(stat.get('bk_percentages', {}).keys())
    bk_codes = sorted(list(bk_codes))

    # Column headers
    headers = ["Leerling"] + bk_codes + ["Totaal %"]
    ws.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Data rows
    for row_idx, (leerling, stat) in enumerate(zip(leerlingen, stats), 4):
        naam = f"{leerling.get('voornaam', '')} {leerling.get('naam', '')}".strip()
        rij = [naam]

        # BK percentages
        bk_percentages = stat.get('bk_percentages', {})
        totaal_pct = stat.get('totaal_percentage', 0)

        for bk_code in bk_codes:
            pct = bk_percentages.get(bk_code, 0)
            if pct >= 100:
                rij.append("100%")
            elif pct > 0:
                rij.append(f"{pct:.0f}%")
            else:
                rij.append("—")

        # Totaal
        rij.append(f"{totaal_pct:.1f}%")

        ws.append(rij)

        # Format cells
        is_even = (row_idx - 4) % 2 == 0
        for col_idx, val in enumerate(rij, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            if col_idx == 1:
                # Leerling name
                cell.alignment = left_align
            else:
                # Percentage cells
                cell.alignment = center_align

                if is_even:
                    cell.fill = alt_fill

                # Color based on percentage
                try:
                    if "%" in str(val):
                        pct_val = float(str(val).replace("%", ""))
                        if pct_val >= 100:
                            cell.fill = behaald_fill
                            cell.font = behaald_font
                        elif pct_val >= 50:
                            cell.fill = gedeeltelijk_fill
                        else:
                            cell.fill = niet_behaald_fill
                except ValueError:
                    pass

    # Column widths
    ws.column_dimensions["A"].width = 25
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 12

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _genereer_csv_bytes(leerling: dict, doelen: list, resultaten: dict) -> bytes:
    """Fallback CSV voor Excel attestering."""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output, delimiter=";")

    naam = f"{leerling.get('voornaam', '')} {leerling.get('naam', '')}".strip()
    schooljaar = datetime.now().year

    writer.writerow([f"Attestering: {naam}"])
    writer.writerow([f"Schooljaar: {schooljaar - 1}-{schooljaar}"])
    writer.writerow([])

    writer.writerow(["Code", "Omschrijving", "Soort", "Behaald", "Datum"])

    for doel in doelen:
        bk_code = doel.get('bk_code', doel.get('code', ''))
        lpd_id = doel.get('lpd_id', doel.get('id'))
        behaald = resultaten.get(lpd_id, False)
        writer.writerow([
            bk_code,
            doel.get('lpd_omschrijving', doel.get('omschrijving', '')),
            "LPD",
            "JA" if behaald else "NEE",
            ""
        ])

    return output.getvalue().encode("utf-8-sig")


def _genereer_csv_bytes_klas(klas: dict, leerlingen: list, stats: list) -> bytes:
    """Fallback CSV voor Excel klasoverzicht."""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output, delimiter=";")

    klas_naam = klas.get('naam', f"Klas {klas.get('id')}")
    schooljaar = klas.get('schooljaar', datetime.now().year)

    writer.writerow([f"Klasoverzicht: {klas_naam}"])
    writer.writerow([f"Schooljaar: {schooljaar - 1}-{schooljaar}"])
    writer.writerow([])

    # Extract unique BK codes
    bk_codes = set()
    for stat in stats:
        bk_codes.update(stat.get('bk_percentages', {}).keys())
    bk_codes = sorted(list(bk_codes))

    # Headers
    headers = ["Leerling"] + bk_codes + ["Totaal %"]
    writer.writerow(headers)

    # Data
    for leerling, stat in zip(leerlingen, stats):
        naam = f"{leerling.get('voornaam', '')} {leerling.get('naam', '')}".strip()
        rij = [naam]

        bk_percentages = stat.get('bk_percentages', {})
        for bk_code in bk_codes:
            pct = bk_percentages.get(bk_code, 0)
            rij.append(f"{pct:.0f}%")

        rij.append(f"{stat.get('totaal_percentage', 0):.1f}%")
        writer.writerow(rij)

    return output.getvalue().encode("utf-8-sig")
